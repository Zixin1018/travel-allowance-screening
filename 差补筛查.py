#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
差补重复筛查工具 v3.0
功能：从月度差补核对报表中自动筛查两类问题 ——
  1. 单流程金额偏大（人工计算天数×标准时出错）
  2. 重叠重复发放（同一员工多流程日期重叠）
联动境外差旅补贴计算器，自动识别境外流程并用对应标准+汇率计算。

输入：差补核对报表.xlsx（含主表数据和明细表数据）
输出：差补筛查结果-{月份}.xlsx（4个Sheet）

用法：python 差补筛查.py 报表文件.xlsx
"""

import os
import sys
import re
import json
from datetime import datetime, date, timedelta
from collections import defaultdict, Counter
from pathlib import Path

# ========== 依赖检查 ==========
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)

# ========== 常量 ==========
STD_FOOD = 100    # 境内伙食补助标准（元/天）
STD_TRAFFIC = 80  # 境内交通补助标准（元/天）
OVERSEAS_PREFIXES = ['亚洲-', '欧洲-', '美洲-', '非洲-', '大洋洲-']


# =============================================================================
# 境外标准 + 汇率 提取（从境外补贴计算器 HTML）
# =============================================================================

def load_overseas_standards(html_path=None):
    """从差旅补贴计算.html提取273条城市标准"""
    if html_path is None:
        html_path = os.path.expanduser('~/dscc产出/境外补贴计算器/差旅补贴计算.html')

    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()

    match = re.search(r'\[{"country":.*?}\](?=\s*[;,\n])', html, re.DOTALL)
    if not match:
        print("[WARN] 未找到境外城市标准数据")
        return {}

    cities = json.loads(match.group())
    # 构建查找表：{国家: [(city_str, currency, food, public), ...]}
    lookup = defaultdict(list)
    for c in cities:
        lookup[c['country']].append({
            'city': c.get('city', ''),
            'currency': c['currency'],
            'food': c['food'],
            'public': c['public'],
        })
    print(f"[OK] 加载境外标准: {len(cities)} 条, {len(lookup)} 个国家/地区")
    return lookup


def load_exchange_rates(html_path=None):
    """从差旅补贴计算.html提取嵌入式汇率数据"""
    if html_path is None:
        html_path = os.path.expanduser('~/dscc产出/境外补贴计算器/差旅补贴计算.html')

    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()

    match = re.search(r'<script type="application/json" id="embeddedRatesData">(.*?)</script>', html, re.DOTALL)
    if not match:
        print("[WARN] 未找到嵌入式汇率数据，境外流程将无法换算")
        return {}

    rates = json.loads(match.group(1))
    print(f"[OK] 加载汇率: {len(rates)} 天, 币种: USD EUR JPY HKD GBP")
    return rates


def get_rate_for_date(rates, date_val, currency):
    """获取指定日期的汇率，无数据时前后各找30天取最近值"""
    if not rates:
        return None

    date_str = date_val.strftime('%Y-%m-%d') if isinstance(date_val, (date, datetime)) else str(date_val)[:10]

    if date_str in rates and currency in rates[date_str]:
        return rates[date_str][currency]

    # 前后各找30天
    base = date_val if isinstance(date_val, date) else datetime.strptime(date_str, '%Y-%m-%d').date()
    for delta in range(1, 31):
        for d in [base - timedelta(days=delta), base + timedelta(days=delta)]:
            ds = d.strftime('%Y-%m-%d')
            if ds in rates and currency in rates[ds]:
                return rates[ds][currency]

    return None


def is_overseas(dest_str):
    """判断目的地是否为境外"""
    if not dest_str:
        return False
    for prefix in OVERSEAS_PREFIXES:
        if dest_str.startswith(prefix):
            return True
    return False


def parse_overseas_location(dest_str):
    """解析境外目的地 → (国家, 城市)"""
    if not dest_str:
        return None, None
    parts = dest_str.split('-')
    # 格式：大洲-国家-城市 或 大洲-地区名
    if len(parts) >= 3:
        return parts[1], parts[2]
    elif len(parts) == 2:
        return parts[1], ''
    return None, None


def match_overseas_standard(standards_lookup, country, city):
    """匹配境外标准 → {currency, food, public} 或 None"""
    if country not in standards_lookup:
        return None

    entries = standards_lookup[country]
    # 精确城市匹配
    for e in entries:
        if e['city'] and city and city in e['city']:
            return e
    # 模糊匹配
    for e in entries:
        if e['city'] and city:
            # 检查city是否在标准城市列表中的一部分
            for std_city in e['city'].split('、'):
                if city in std_city or std_city in city:
                    return e
    # 无城市匹配 → 用"其他城市"或第一个
    for e in entries:
        if not e['city'] or '其他' in e['city']:
            return e
    return entries[0] if entries else None


# =============================================================================
# 数据加载
# =============================================================================

def scan_headers(ws, header_row=1):
    """扫描表头行，返回 {列名(大写): 列索引(0-based)}"""
    header_map = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=col).value
        if h and isinstance(h, str):
            header_map[h.upper().strip()] = col - 1
    return header_map


def get_col(header_map, col_name):
    """从表头映射获取列索引，精确+模糊匹配，找不到返回None"""
    if not col_name:
        return None
    key = col_name.upper()
    if key in header_map:
        return header_map[key]
    for h, idx in header_map.items():
        if key in h or h in key:
            return idx
    return None


def load_data(file_path):
    """加载主表数据和明细表数据"""
    wb = load_workbook(file_path, data_only=True)

    # --- 主表数据 ---
    print("[INFO] 读取主表数据...")
    ws_main = wb['主表数据']
    hm = scan_headers(ws_main)

    col_proc = get_col(hm, '流程编号')
    col_emp_id = get_col(hm, '工号')
    col_emp_name = get_col(hm, '出差人')
    col_time = get_col(hm, '起止时间')
    col_allowance = get_col(hm, '差补')
    col_synced = get_col(hm, '同步资金系统')

    # 校验必要列
    required = {'流程编号': col_proc, '工号': col_emp_id, '出差人': col_emp_name,
                '起止时间': col_time, '差补': col_allowance}
    missing = [k for k, v in required.items() if v is None]
    if missing:
        print(f"[ERR] 主表数据缺少列: {missing}")
        wb.close()
        return [], {}, {}

    main_data = []
    for row in ws_main.iter_rows(min_row=2, values_only=True):
        proc_id = str(row[col_proc]).strip() if row[col_proc] else None
        if not proc_id:
            continue
        emp_id = str(row[col_emp_id]).strip() if row[col_emp_id] else ''
        emp_name = str(row[col_emp_name]).strip() if row[col_emp_name] else ''
        time_range = str(row[col_time]).strip() if row[col_time] else ''
        allowance = float(row[col_allowance]) if row[col_allowance] else 0

        # 解析起止日期
        start_date, end_date = None, None
        if '~' in time_range:
            parts = time_range.split('~')
            for fmt in ('%Y-%m-%d', '%Y/%m/%d'):
                try:
                    start_date = datetime.strptime(parts[0].strip(), fmt).date()
                    end_date = datetime.strptime(parts[1].strip(), fmt).date()
                    break
                except ValueError:
                    continue

        if not start_date or not end_date:
            continue

        synced = False
        if col_synced is not None:
            v = str(row[col_synced]).strip() if row[col_synced] else ''
            synced = '已同步' in v
        main_data.append({
            'proc_id': proc_id,
            'emp_id': emp_id,
            'emp_name': emp_name,
            'start_date': start_date,
            'end_date': end_date,
            'allowance': allowance,
            'synced': synced,
        })

    print(f"  主表: {len(main_data)} 条")

    # --- 明细表数据 ---
    print("[INFO] 读取明细表数据...")
    ws_detail = wb['明细表数据']
    hd = scan_headers(ws_detail)

    col_d_proc = get_col(hd, '流程编号')
    col_d_emp_id = get_col(hd, '工号')
    col_d_emp_name = get_col(hd, '出差人')
    col_d_time = get_col(hd, '起止时间')
    col_d_dep = get_col(hd, '出发地')
    col_d_dest = get_col(hd, '目的地')
    col_d_food = get_col(hd, '伙食补助')
    col_d_traffic = get_col(hd, '交通补助')
    col_d_note = get_col(hd, '差补说明')

    required_d = {'流程编号': col_d_proc, '出发地': col_d_dep, '目的地': col_d_dest,
                  '伙食补助': col_d_food, '交通补助': col_d_traffic}
    missing_d = [k for k, v in required_d.items() if v is None]
    if missing_d:
        print(f"[ERR] 明细表数据缺少列: {missing_d}")
        wb.close()
        return main_data, {}, {}

    detail_by_flow = defaultdict(lambda: {
        'food_sum': 0, 'traffic_sum': 0,
        'min_date': None, 'max_date': None,
        'destinations': set(), 'departures': set(),
        'notes': set(),
    })

    detail_rows = 0
    skipped_zero = 0
    detail_emp_info = {}
    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        proc_id = str(row[col_d_proc]).strip() if row[col_d_proc] else None
        if not proc_id:
            continue

        food = float(row[col_d_food]) if row[col_d_food] else 0
        traffic = float(row[col_d_traffic]) if row[col_d_traffic] else 0
        departure = str(row[col_d_dep]).strip() if row[col_d_dep] else ''
        destination = str(row[col_d_dest]).strip() if row[col_d_dest] else ''

        if col_d_emp_id is not None:
            emp_id = str(row[col_d_emp_id]).strip() if row[col_d_emp_id] else ''
            emp_name = str(row[col_d_emp_name]).strip() if row[col_d_emp_name] and col_d_emp_name is not None else ''
            if emp_id and proc_id not in detail_emp_info:
                detail_emp_info[proc_id] = (emp_id, emp_name)

        time_range = str(row[col_d_time]).strip() if col_d_time is not None and row[col_d_time] else ''
        seg_start, seg_end = None, None
        if '~' in time_range:
            parts = time_range.split('~')
            for fmt in ('%Y-%m-%d', '%Y/%m/%d'):
                try:
                    seg_start = datetime.strptime(parts[0].strip(), fmt).date()
                    seg_end = datetime.strptime(parts[1].strip(), fmt).date()
                    break
                except ValueError:
                    continue

        # 跳过金额全为0的行
        if food == 0 and traffic == 0:
            skipped_zero += 1
            continue

        detail_rows += 1
        d = detail_by_flow[proc_id]
        d['food_sum'] += food
        d['traffic_sum'] += traffic
        if seg_start:
            if d['min_date'] is None or seg_start < d['min_date']:
                d['min_date'] = seg_start
        if seg_end:
            if d['max_date'] is None or seg_end > d['max_date']:
                d['max_date'] = seg_end
        if destination:
            d['destinations'].add(destination)
        if departure:
            d['departures'].add(departure)
        note = str(row[col_d_note]).strip() if col_d_note is not None and row[col_d_note] else ''
        if note:
            d['notes'].add(note)

    print(f"  明细: {detail_rows} 条有效行, 跳过 {skipped_zero} 条零金额")
    wb.close()
    return main_data, dict(detail_by_flow), detail_emp_info


# =============================================================================
# 数据聚合（构建差补重复分析）
# =============================================================================

def build_analysis_data(main_data, detail_by_flow, standards_lookup, detail_emp_info=None):
    """合并主表+明细表，构建差补重复分析数据结构"""
    records = []
    seen_procs = set()
    skipped = 0

    for m in main_data:
        pid = m['proc_id']
        seen_procs.add(pid)
        detail = detail_by_flow.get(pid)

        if detail:
            food_amt = detail['food_sum']
            traffic_amt = detail['traffic_sum']
            start_date = detail.get('min_date') or m['start_date']
            end_date = detail.get('max_date') or m['end_date']
            dests = detail['destinations']  # 只看目的地，不看出发地
        else:
            food_amt = m['allowance']
            traffic_amt = 0
            start_date = m['start_date']
            end_date = m['end_date']
            dests = set()

        if food_amt == 0 and traffic_amt == 0:
            skipped += 1
            continue

        days = (end_date - start_date).days + 1
        is_ovs, overseas_info, _ = _detect_overseas(dests, standards_lookup)
        records.append(_make_record(pid, m['emp_id'], m['emp_name'],
                                    start_date, end_date, days, food_amt, traffic_amt,
                                    is_ovs, overseas_info, m.get('synced', False),
                                    detail.get('destinations') if detail else None,
                                    detail.get('departures') if detail else None,
                                    detail.get('notes') if detail else None))

    # 补充：仅明细表有但主表无的流程（非零金额）
    if detail_emp_info:
        extra_count = 0
        for pid, detail in detail_by_flow.items():
            if pid in seen_procs:
                continue
            food_amt = detail['food_sum']
            traffic_amt = detail['traffic_sum']
            if food_amt == 0 and traffic_amt == 0:
                continue
            start_date = detail.get('min_date')
            end_date = detail.get('max_date')
            if not start_date or not end_date:
                continue

            emp_id, emp_name = detail_emp_info.get(pid, ('', ''))
            dests = detail['destinations']
            days = (end_date - start_date).days + 1
            is_ovs, overseas_info, _ = _detect_overseas(dests, standards_lookup)
            records.append(_make_record(pid, emp_id, emp_name,
                                        start_date, end_date, days, food_amt, traffic_amt,
                                        is_ovs, overseas_info, False,
                                        detail['destinations'], detail['departures'],
                                        detail['notes']))
            extra_count += 1
        if extra_count:
            print(f"[INFO] 补充仅明细有的流程: {extra_count} 条")

    print(f"[INFO] 差补重复分析: {len(records)} 条, 跳过零金额 {skipped} 条")
    overseas_count = sum(1 for r in records if r['is_overseas'])
    print(f"[INFO] 其中境外流程: {overseas_count} 条")
    return records


def _detect_overseas(destinations, standards_lookup):
    """检测是否为境外（只看目的地），返回(is_overseas, overseas_info, all_countries)"""
    overseas_dests = []
    for dest in sorted(destinations):  # 排序保证确定性
        if is_overseas(dest):
            overseas_dests.append(dest)

    if not overseas_dests:
        return False, None, []

    # 收集所有境外国家/地区
    countries = []
    for dest in overseas_dests:
        country, city = parse_overseas_location(dest)
        if country and country not in countries:
            countries.append(country)

    # 用第一个匹配到的做标准（主目的地）
    for dest in overseas_dests:
        country, city = parse_overseas_location(dest)
        if country:
            std = match_overseas_standard(standards_lookup, country, city)
            if std:
                return True, {
                    'country': '、'.join(countries) if len(countries) > 1 else country,
                    'city': city,
                    'currency': std['currency'],
                    'food_std': std['food'], 'public_std': std['public'],
                    'dest_raw': dest,
                    'all_countries': countries,
                }, countries

    return False, None, []


def _make_record(pid, emp_id, emp_name, start_date, end_date, days, food_amt, traffic_amt,
                 is_ovs, overseas_info, synced=False, destinations=None, departures=None, notes=None):
    return {
        'proc_id': pid, 'emp_id': emp_id, 'emp_name': emp_name,
        'start_date': start_date, 'end_date': end_date, 'days': days,
        'food_amt': food_amt, 'traffic_amt': traffic_amt,
        'total_amt': food_amt + traffic_amt,
        'is_overseas': is_ovs, 'overseas_info': overseas_info,
        'synced': synced,
        'destinations': '、'.join(sorted(destinations)) if destinations else '',
        'departures': '、'.join(sorted(departures)) if departures else '',
        'notes': '; '.join(sorted(notes)) if notes else '',
    }


# =============================================================================
# 境外流程计算
# =============================================================================

def calc_overseas_expected(records, rates, standards_lookup):
    """为境外流程计算应有金额（标准×汇率），并检测境外流程间的重叠"""
    overseas_results = []

    # 先按员工分组，检测境外流程间的重叠
    overseas_by_emp = defaultdict(list)
    for r in records:
        if r['is_overseas'] and r['overseas_info']:
            overseas_by_emp[(r['emp_id'], r['emp_name'])].append(r)

    overseas_overlap_procs = set()
    for emp_key, emp_recs in overseas_by_emp.items():
        if len(emp_recs) <= 1:
            continue
        blocks = group_overlap_blocks(emp_recs)
        for block in blocks:
            if len(block) > 1:
                date_counter = Counter()
                for r in block:
                    current = r['start_date']
                    while current <= r['end_date']:
                        date_counter[current] += 1
                        current += timedelta(days=1)
                overlap_dates = sorted([d for d, cnt in date_counter.items() if cnt > 1])
                if overlap_dates:
                    for r in block:
                        overseas_overlap_procs.add(r['proc_id'])

    # 逐流程计算应有金额（按返程日期汇率，合规要求）
    for r in records:
        if not r['is_overseas'] or not r['overseas_info']:
            continue

        info = r['overseas_info']
        currency = info['currency']
        food_std = info['food_std']
        public_std = info['public_std']
        days = r['days']

        # 用返程日期（结束日期）的汇率，整段行程统一汇率
        rate = get_rate_for_date(rates, r['end_date'], currency)
        if rate:
            total_expected_food = round(food_std * days * rate, 2)
            total_expected_traffic = round(public_std * days * rate, 2)
            missing_rate = False
        else:
            total_expected_food = None
            total_expected_traffic = None
            missing_rate = True

        food_diff = r['food_amt'] - total_expected_food if total_expected_food is not None else None
        traffic_diff = r['traffic_amt'] - total_expected_traffic if total_expected_traffic is not None else None

        # 预警标记（只标多发和重叠）
        warnings = []
        if missing_rate:
            warnings.append('返程日期无汇率')
        if r['proc_id'] in overseas_overlap_procs:
            warnings.append('境外流程日期重叠')
        if info.get('all_countries') and len(info['all_countries']) > 1:
            warnings.append(f"涉及多国({','.join(info['all_countries'])})，仅用首国标准计算")
        if food_diff is not None and food_diff > 50:
            warnings.append('伙食多发')
        if traffic_diff is not None and traffic_diff > 50:
            warnings.append('交通多发')

        overseas_results.append({
            **r,
            'expected_food': round(total_expected_food, 2) if total_expected_food is not None else None,
            'expected_traffic': round(total_expected_traffic, 2) if total_expected_traffic is not None else None,
            'food_diff': round(food_diff, 2) if food_diff is not None else None,
            'traffic_diff': round(traffic_diff, 2) if traffic_diff is not None else None,
            'return_rate': rate,
            'has_overlap': r['proc_id'] in overseas_overlap_procs,
            'warnings': '; '.join(warnings) if warnings else '',
        })

    return overseas_results


# =============================================================================
# 筛查：单流程金额检查（境内）
# =============================================================================

def check_single_flow(records):
    """检查单流程金额是否偏大（仅境内）"""
    issues = []
    for r in records:
        if r['is_overseas']:
            continue

        expected_food = r['days'] * STD_FOOD
        expected_traffic = r['days'] * STD_TRAFFIC
        food_diff = r['food_amt'] - expected_food
        traffic_diff = r['traffic_amt'] - expected_traffic

        if food_diff > 1 or traffic_diff > 1:
            desc_parts = []
            if food_diff > 1:
                desc_parts.append(f"伙食实发{r['food_amt']:.0f}元,应有{expected_food:.0f}元,多发{food_diff:.0f}元")
            if traffic_diff > 1:
                desc_parts.append(f"交通实发{r['traffic_amt']:.0f}元,应有{expected_traffic:.0f}元,多发{traffic_diff:.0f}元")

            issues.append({
                **r,
                'expected_food': expected_food,
                'expected_traffic': expected_traffic,
                'food_diff': food_diff,
                'traffic_diff': traffic_diff,
                'issue_type': '单流程金额偏大',
                'desc': '; '.join(desc_parts),
            })

    return issues


# =============================================================================
# 筛查：重叠重复发放（境内）
# =============================================================================

def group_overlap_blocks(records):
    """将同一员工的流程按连通重叠块分组"""
    if not records:
        return []
    sorted_recs = sorted(records, key=lambda r: r['start_date'])
    n = len(sorted_recs)
    block_ids = list(range(n))

    for i in range(1, n):
        for j in range(i):
            if (sorted_recs[i]['start_date'] <= sorted_recs[j]['end_date'] and
                sorted_recs[j]['start_date'] <= sorted_recs[i]['end_date']):
                old_id = block_ids[i]
                new_id = block_ids[j]
                for k in range(n):
                    if block_ids[k] == old_id:
                        block_ids[k] = new_id

    blocks = defaultdict(list)
    for i, rec in enumerate(sorted_recs):
        blocks[block_ids[i]].append(rec)
    return list(blocks.values())


def expand_dates(records):
    """展开流程列表中的所有日期（去重）"""
    date_set = set()
    for r in records:
        current = r['start_date']
        while current <= r['end_date']:
            date_set.add(current)
            current += timedelta(days=1)
    return sorted(date_set)


def format_date_range(dates):
    """日期列表 → 可读范围字符串"""
    if not dates:
        return ''
    parts = []
    i = 0
    while i < len(dates):
        start = dates[i]
        end = start
        while i + 1 < len(dates) and (dates[i + 1] - dates[i]).days == 1:
            i += 1
            end = dates[i]
        if start == end:
            parts.append(start.strftime('%m/%d'))
        else:
            parts.append(f"{start.strftime('%m/%d')}~{end.strftime('%m/%d')}")
        i += 1
    return '、'.join(parts)


def check_overlap(records, rates):
    """检查重叠重复发放（境内外统一处理，境外优先）"""
    # 按员工分组（不区分境内外）
    emp_groups = defaultdict(list)
    for r in records:
        emp_groups[(r['emp_id'], r['emp_name'])].append(r)

    overlap_issues = []
    notable_no_overpay = []  # 日期重叠但无多发（标绿用）
    emp_summary = []

    for (emp_id, emp_name), emp_records in emp_groups.items():
        # 分离境内境外
        domestic = [r for r in emp_records if not r['is_overseas']]
        overseas = [r for r in emp_records if r['is_overseas']]

        # 构建日期覆盖图：每天标记是否境外、被哪些流程覆盖
        date_map = {}  # date -> {'is_overseas': bool, 'flows': set, 'overseas_info': dict}
        for r in emp_records:
            current = r['start_date']
            while current <= r['end_date']:
                if current not in date_map:
                    date_map[current] = {'is_overseas': False, 'flows': set()}
                if r['is_overseas']:
                    date_map[current]['is_overseas'] = True
                    if r['overseas_info']:
                        date_map[current]['overseas_info'] = r['overseas_info']
                date_map[current]['flows'].add(r['proc_id'])
                current += timedelta(days=1)

        # 统计境外天和境内天（重叠日境外优先）
        overseas_dates = sorted([d for d, m in date_map.items() if m['is_overseas']])
        domestic_dates = sorted([d for d, m in date_map.items() if not m['is_overseas']])
        overlap_dates = sorted([d for d, m in date_map.items() if len(m['flows']) > 1])

        ovs_days = len(overseas_dates)
        dom_days = len(domestic_dates)

        # 应有金额
        # 境外：逐日按各自标准+汇率计算
        exp_food_ovs = 0
        exp_traffic_ovs = 0
        has_missing_rate = False
        for d in overseas_dates:
            info = date_map[d].get('overseas_info')
            if info:
                rate = get_rate_for_date(rates, d, info['currency'])
                if rate:
                    exp_food_ovs += info['food_std'] * rate
                    exp_traffic_ovs += info['public_std'] * rate
                else:
                    has_missing_rate = True

        # 境内：100/80 × 境内天
        exp_food_dom = dom_days * STD_FOOD
        exp_traffic_dom = dom_days * STD_TRAFFIC

        total_exp_food = exp_food_ovs + exp_food_dom
        total_exp_traffic = exp_traffic_ovs + exp_traffic_dom

        # 实发合计
        actual_food = sum(r['food_amt'] for r in emp_records)
        actual_traffic = sum(r['traffic_amt'] for r in emp_records)

        food_over = actual_food - total_exp_food
        traffic_over = actual_traffic - total_exp_traffic

        # 写入汇总
        emp_summary.append({
            'emp_id': emp_id, 'emp_name': emp_name,
            'proc_count': len(emp_records),
            'overseas_days': ovs_days, 'domestic_days': dom_days,
            'total_food': actual_food, 'exp_food': round(total_exp_food, 2),
            'food_over': round(max(0, food_over), 2),
            'total_traffic': actual_traffic, 'exp_traffic': round(total_exp_traffic, 2),
            'traffic_over': round(max(0, traffic_over), 2),
        })

        # 计算参与重叠的流程
        overlap_procs = set()
        for d in overlap_dates:
            overlap_procs |= date_map[d]['flows']

        # 有重叠且多发 → 生成明细
        if overlap_dates and (food_over > 1 or traffic_over > 1):
            overlap_desc = format_date_range(overlap_dates)
            desc_parts = [f"重叠日期: {overlap_desc}"]
            mixed_overlap = any(
                date_map[d]['is_overseas'] and
                any(not r2['is_overseas'] for r2 in emp_records if r2['proc_id'] in date_map[d]['flows'])
                for d in overlap_dates
            )
            if ovs_days > 0:
                note = '(优先)' if mixed_overlap else ''
                desc_parts.append(f"境外{ovs_days}天{note}+境内{dom_days}天")
            else:
                desc_parts.append(f"合并去重{dom_days}天")
            if food_over > 1:
                desc_parts.append(f"伙食:实发{actual_food:.0f},应有{total_exp_food:.0f},多发{food_over:.0f}")
            if traffic_over > 1:
                desc_parts.append(f"交通:实发{actual_traffic:.0f},应有{total_exp_traffic:.0f},多发{traffic_over:.0f}")

            for r in emp_records:
                if r['proc_id'] not in overlap_procs:
                    continue
                if r['is_overseas']:
                    exp_f = r.get('expected_food') or 0
                    exp_t = r.get('expected_traffic') or 0
                else:
                    exp_f = r['days'] * STD_FOOD
                    exp_t = r['days'] * STD_TRAFFIC

                overlap_issues.append({
                    **r,
                    'expected_food': exp_f,
                    'expected_traffic': exp_t,
                    'food_diff': r['food_amt'] - exp_f,
                    'traffic_diff': r['traffic_amt'] - exp_t,
                    'issue_type': '重复发放',
                    'desc': '; '.join(desc_parts),
                    'overlap_dates': overlap_desc,
                    'overlap_proc_count': len(emp_records),
                })
        elif overlap_dates:
            # 有重叠但无多发 → 标绿提醒
            for proc_id in overlap_procs:
                notable_no_overpay.append(proc_id)

    return overlap_issues, emp_summary, notable_no_overpay


# =============================================================================
# 生成输出 Excel
# =============================================================================

def create_output(records, single_issues, overlap_issues, emp_summary, overseas_results, notable_no_overpay, input_file, output_path):
    """生成筛查结果 Excel"""
    wb = Workbook()

    # 样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_font = Font(color='FF0000', bold=True, size=10)
    blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    gray_font = Font(color='808080', size=10)
    normal_font = Font(size=10)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 境外需关注
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 分类标色：黄=确认多发，绿=需关注（重叠/多国等但无多发）
    yellow_proc_ids = set()  # 确认多发
    green_proc_ids = set()   # 需关注
    for r in single_issues:
        yellow_proc_ids.add(r['proc_id'])
    for r in overlap_issues:
        yellow_proc_ids.add(r['proc_id'])
    for proc_id in notable_no_overpay:
        green_proc_ids.add(proc_id)
    for r in overseas_results:
        has_overlap = r.get('has_overlap')
        has_overpay = r.get('warnings') and '多发' in str(r.get('warnings', ''))
        has_warning = bool(r.get('warnings'))
        if has_overpay:
            yellow_proc_ids.add(r['proc_id'])
        elif has_overlap or has_warning:
            green_proc_ids.add(r['proc_id'])

    # ===== Sheet 1: 差补筛查明细 =====
    ws1 = wb.active
    ws1.title = '差补筛查明细'

    headers1 = [
        '流程编号', '工号', '姓名', '流程起始时间', '流程结束时间',
        '伙食补助', '交通补助', '伙食+交通补助',
        '类型', '天数', '伙食应有金额', '交通应有金额',
        '伙食差额', '交通差额', '说明'
    ]
    for ci, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = header_fill, header_font, center_align, thin_border

    # 合并境内异常 + 境外重叠/多发
    all_issues = single_issues + overlap_issues
    for r in overseas_results:
        has_overlap = r.get('has_overlap')
        has_overpay = r.get('warnings') and '多发' in str(r.get('warnings', ''))
        if has_overlap or has_overpay:
            exp_f = r.get('expected_food') or 0
            exp_t = r.get('expected_traffic') or 0
            all_issues.append({
                **r,
                'expected_food': exp_f, 'expected_traffic': exp_t,
                'food_diff': round(r['food_amt'] - exp_f, 2) if exp_f else None,
                'traffic_diff': round(r['traffic_amt'] - exp_t, 2) if exp_t else None,
                'issue_type': '境外日期重叠' if has_overlap else '境外金额多发',
                'desc': r.get('warnings', ''),
            })
    all_issues.sort(key=lambda r: (r['emp_id'], r['start_date']))

    for ri, r in enumerate(all_issues, 2):
        vals = [
            r['proc_id'], r['emp_id'], r['emp_name'],
            r['start_date'], r['end_date'],
            r['food_amt'], r['traffic_amt'], r['total_amt'],
            r['issue_type'], r['days'],
            r['expected_food'], r['expected_traffic'],
            r['food_diff'], r['traffic_diff'], r['desc'],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=v)
            c.font, c.border = normal_font, thin_border
            c.alignment = center_align if ci in (1, 2, 9, 10) else left_align
        is_green = r['proc_id'] in green_proc_ids
        for c in range(1, len(headers1) + 1):
            ws1.cell(row=ri, column=c).fill = green_fill if is_green else yellow_fill

    col_widths1 = {1: 30, 2: 12, 3: 12, 4: 14, 5: 14, 6: 12, 7: 12, 8: 14,
                   9: 16, 10: 8, 11: 14, 12: 14, 13: 16, 14: 16, 15: 80}
    for col, w in col_widths1.items():
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = 'A2'

    # ===== Sheet 2: 原始数据(标黄) =====
    ws2 = wb.create_sheet('原始数据(标黄)')

    headers2 = ['流程编号', '工号', '姓名', '流程起始时间', '流程结束时间',
                '天数', '伙食补助', '交通补助', '合计', '境内外',
                '同步状态', '出发地', '目的地', '差补说明',
                '境外国家', '境外币种', '境外伙食标准(原币)', '境外公杂标准(原币)',
                '类型标记']
    for ci, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = header_fill, header_font, center_align, thin_border

    records_sorted = sorted(records, key=lambda r: (r['emp_id'], r['start_date']))
    for ri, r in enumerate(records_sorted, 2):
        is_synced = r.get('synced', False)
        issue_type = ''
        proc_fill = None
        row_font = gray_font if is_synced else normal_font

        if not is_synced:
            if r['proc_id'] in yellow_proc_ids:
                types = []
                for si in single_issues:
                    if si['proc_id'] == r['proc_id']:
                        types.append('单流程金额偏大')
                        break
                for oi in overlap_issues:
                    if oi['proc_id'] == r['proc_id']:
                        types.append('重复发放')
                        break
                issue_type = ';'.join(types)
                proc_fill = yellow_fill
            elif r['proc_id'] in green_proc_ids:
                issue_type = '境外需关注'
                proc_fill = green_fill
        elif r['proc_id'] in yellow_proc_ids or r['proc_id'] in green_proc_ids:
            issue_type = '已同步-不可修改'

        ovs = r.get('overseas_info') or {}
        vals = [
            r['proc_id'], r['emp_id'], r['emp_name'],
            r['start_date'], r['end_date'], r['days'],
            r['food_amt'], r['traffic_amt'], r['total_amt'],
            '境外' if r['is_overseas'] else '境内',
            '已同步' if is_synced else '未同步',
            r.get('departures', ''),
            r.get('destinations', ''),
            r.get('notes', ''),
            ovs.get('country', ''),
            ovs.get('currency', ''),
            ovs.get('food_std', ''),
            ovs.get('public_std', ''),
            issue_type,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = row_font
            c.border = thin_border
            c.alignment = center_align if ci in (2, 6, 10, 11) else left_align
            if r['is_overseas'] and ci == 10:
                c.font = red_font
            if is_synced:
                c.font = gray_font

        if proc_fill:
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=ri, column=c).fill = proc_fill

    col_widths2 = {1: 30, 2: 12, 3: 12, 4: 14, 5: 14, 6: 8, 7: 12, 8: 12,
                   9: 12, 10: 8, 11: 10, 12: 20, 13: 20, 14: 30, 15: 12, 16: 10,
                   17: 18, 18: 18, 19: 20}
    for col, w in col_widths2.items():
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A2'

    # ===== Sheet 3: 汇总统计 =====
    ws3 = wb.create_sheet('汇总统计')

    headers3 = ['工号', '姓名', '涉及流程数', '境外天', '境内天', '去重总天数',
                '伙食实发合计', '伙食应有合计', '伙食多发金额',
                '交通实发合计', '交通应有合计', '交通多发金额']
    for ci, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = header_fill, header_font, center_align, thin_border

    emp_summary.sort(key=lambda r: (r['food_over'] + r['traffic_over']), reverse=True)

    for ri, r in enumerate(emp_summary, 2):
        unique_total = r.get('overseas_days', 0) + r.get('domestic_days', 0)
        vals = [r['emp_id'], r['emp_name'], r['proc_count'],
                r.get('overseas_days', 0), r.get('domestic_days', 0), unique_total,
                r['total_food'], r['exp_food'], r['food_over'],
                r['total_traffic'], r['exp_traffic'], r['traffic_over']]
        for ci, v in enumerate(vals, 1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.font, c.border = normal_font, thin_border
            c.alignment = center_align
            if ci == 9 and v > 0:
                c.font = red_font
            if ci == 12 and v > 0:
                c.font = red_font

    col_widths3 = {1: 14, 2: 12, 3: 12, 4: 8, 5: 8, 6: 12, 7: 14, 8: 14, 9: 14, 10: 14, 11: 14, 12: 14}
    for col, w in col_widths3.items():
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = 'A2'

    # ===== Sheet 4: 境外流程清单 =====
    ws4 = wb.create_sheet('境外流程清单')

    headers4 = ['流程编号', '工号', '姓名', '起始日期', '结束日期', '天数',
                '目的地', '国家', '币种', '伙食标准(原币/天)', '公杂标准(原币/天)',
                '伙食实发(RMB)', '伙食应有(RMB)', '伙食差异',
                '交通实发(RMB)', '交通应有(RMB)', '交通差异',
                '返程汇率', '预警标记']
    for ci, h in enumerate(headers4, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = header_fill, header_font, center_align, thin_border

    for ri, r in enumerate(overseas_results, 2):
        info = r.get('overseas_info', {})
        ret_rate = r.get('return_rate')
        rate_status = f"{ret_rate:.6f}" if ret_rate else '无汇率'

        vals = [
            r['proc_id'], r['emp_id'], r['emp_name'],
            r['start_date'], r['end_date'], r['days'],
            info.get('dest_raw', ''),
            info.get('country', ''),
            info.get('currency', ''),
            info.get('food_std', ''),
            info.get('public_std', ''),
            r['food_amt'],
            r.get('expected_food', ''),
            r.get('food_diff', ''),
            r['traffic_amt'],
            r.get('expected_traffic', ''),
            r.get('traffic_diff', ''),
            rate_status,
            r.get('warnings', ''),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.font, c.border = normal_font, thin_border
            c.alignment = center_align if ci in (2, 6, 9, 18) else left_align

        # 差异为正(多发)标红
        food_diff = r.get('food_diff')
        if food_diff is not None and food_diff > 50:
            ws4.cell(row=ri, column=14).font = red_font
        traffic_diff = r.get('traffic_diff')
        if traffic_diff is not None and traffic_diff > 50:
            ws4.cell(row=ri, column=17).font = red_font

        # 境外标色：黄=多发，绿=需关注（重叠/多国等）
        if r['proc_id'] in yellow_proc_ids:
            for c in range(1, len(headers4) + 1):
                ws4.cell(row=ri, column=c).fill = yellow_fill
        elif r['proc_id'] in green_proc_ids:
            for c in range(1, len(headers4) + 1):
                ws4.cell(row=ri, column=c).fill = green_fill

    col_widths4 = {1: 28, 2: 10, 3: 10, 4: 12, 5: 12, 6: 6, 7: 20, 8: 10,
                   9: 6, 10: 16, 11: 16, 12: 14, 13: 14, 14: 14,
                   15: 14, 16: 14, 17: 14, 18: 15, 19: 40}
    for col, w in col_widths4.items():
        ws4.column_dimensions[get_column_letter(col)].width = w
    ws4.freeze_panes = 'A2'

    # ===== Sheet 5 & 6: 原始主表+明细表（附境内外标记）=====
    # 构建流程→境外标记映射
    proc_overseas_map = {r['proc_id']: r['is_overseas'] for r in records}

    # 重新读取原始文件
    try:
        src_wb = load_workbook(input_file, data_only=True)
    except Exception:
        src_wb = None

    if src_wb:
        for src_sheet_name in ['主表数据', '明细表数据']:
            if src_sheet_name not in src_wb.sheetnames:
                continue
            src_ws = src_wb[src_sheet_name]
            dst_ws = wb.create_sheet(f'原始{src_sheet_name}')

            # 复制表头 + 新增"境内外"列
            for c in range(1, src_ws.max_column + 1):
                h = src_ws.cell(row=1, column=c).value
                cell = dst_ws.cell(row=1, column=c, value=h)
                cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center_align, thin_border
            dst_ws.cell(row=1, column=src_ws.max_column + 1, value='境内外')
            dst_ws.cell(row=1, column=src_ws.max_column + 1).fill = header_fill
            dst_ws.cell(row=1, column=src_ws.max_column + 1).font = header_font
            dst_ws.cell(row=1, column=src_ws.max_column + 1).alignment = center_align
            dst_ws.cell(row=1, column=src_ws.max_column + 1).border = thin_border

            # 复制数据行
            # 找同步资金系统列
            src_sync_col = get_col(scan_headers(src_ws), '同步资金系统')
            dst_last_col = src_ws.max_column + 1
            for r_idx, row in enumerate(src_ws.iter_rows(min_row=2, values_only=True), 2):
                proc_id = str(row[0]).strip() if row[0] else ''
                is_ovs = proc_overseas_map.get(proc_id, False)
                is_synced = False
                if src_sync_col is not None:
                    v = str(row[src_sync_col]).strip() if row[src_sync_col] else ''
                    is_synced = '已同步' in v
                row_fill = None if is_synced else None  # 默认不标色
                if not is_synced and proc_id in yellow_proc_ids:
                    row_fill = yellow_fill
                elif not is_synced and proc_id in green_proc_ids:
                    row_fill = green_fill
                row_font = gray_font if is_synced else normal_font
                for c_idx, val in enumerate(row, 1):
                    cell = dst_ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = row_font
                    cell.border = thin_border
                    cell.alignment = left_align
                    if row_fill:
                        cell.fill = row_fill
                # 境内外标记
                mark_cell = dst_ws.cell(row=r_idx, column=dst_last_col,
                                        value='境外' if is_ovs else '境内')
                mark_cell.font = red_font if (is_ovs and not is_synced) else row_font
                mark_cell.border = thin_border
                mark_cell.alignment = center_align
                if row_fill:
                    mark_cell.fill = row_fill

            # 调整列宽
            for c in range(1, src_ws.max_column + 2):
                dst_ws.column_dimensions[get_column_letter(c)].width = 15
            dst_ws.freeze_panes = 'A2'

        src_wb.close()

    # 保存
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb.save(output_path)
    print(f"\n[OK] 输出文件: {output_path}")


# =============================================================================
# 主函数
# =============================================================================

def main():
    print("=" * 60)
    print("  差补重复筛查工具 v3.0")
    print("=" * 60)

    # 参数
    if len(sys.argv) < 2:
        print("用法: python 差补筛查.py <报表文件.xlsx>")
        print("示例: python 差补筛查.py ~/Documents/差补核对报表-v1.xlsx")
        sys.exit(1)

    input_file = os.path.abspath(sys.argv[1])
    if not os.path.exists(input_file):
        print(f"[ERR] 文件不存在: {input_file}")
        sys.exit(1)

    print(f"\n[INFO] 输入文件: {input_file}")

    # 1. 加载境外标准+汇率
    print("\n[1/5] 加载境外标准和汇率...")
    standards = load_overseas_standards()
    rates = load_exchange_rates()

    # 2. 加载数据
    print("\n[2/5] 加载报表数据...")
    main_data, detail_by_flow, detail_emp_info = load_data(input_file)

    # 3. 聚合构建差补重复分析
    print("\n[3/5] 构建差补重复分析...")
    records = build_analysis_data(main_data, detail_by_flow, standards, detail_emp_info)

    # 4. 筛查
    print("\n[4/5] 执行筛查...")
    single_issues = check_single_flow(records)
    overseas_results = calc_overseas_expected(records, rates, standards)

    # 将境外预期值注入 records，供 check_overlap 使用
    overseas_exp = {r['proc_id']: r for r in overseas_results}
    for r in records:
        if r['proc_id'] in overseas_exp:
            oe = overseas_exp[r['proc_id']]
            r['expected_food'] = oe.get('expected_food')
            r['expected_traffic'] = oe.get('expected_traffic')

    overlap_issues, emp_summary, notable_no_overpay = check_overlap(records, rates)

    print(f"  单流程金额偏大: {len(single_issues)} 条")
    print(f"  重叠重复发放:   {len(overlap_issues)} 条")
    print(f"  境外流程:       {len(overseas_results)} 条")
    print(f"  员工汇总:       {len(emp_summary)} 人")

    # 5. 确定月份和输出路径
    # 取所有记录的日期的众数月份
    month_counter = Counter()
    for r in records:
        month_counter[r['start_date'].strftime('%Y年%m月')] += 1
        month_counter[r['end_date'].strftime('%Y年%m月')] += 1
    month_str = month_counter.most_common(1)[0][0] if month_counter else datetime.now().strftime('%Y年%m月')

    output_dir = os.path.dirname(input_file)
    output_file = os.path.join(output_dir, f'差补筛查结果-{month_str}.xlsx')

    print(f"\n[5/5] 生成输出文件...")
    create_output(records, single_issues, overlap_issues, emp_summary, overseas_results, notable_no_overpay, input_file, output_file)

    # 汇总
    print("\n" + "=" * 60)
    print("  筛查完成！")
    print("=" * 60)
    print(f"  输出: {output_file}")
    print(f"  境内单流程偏大: {len(single_issues)} 条")
    print(f"  境内重叠重复:   {len(overlap_issues)} 条")
    print(f"  境外流程(已计算): {len(overseas_results)} 条")
    print(f"  员工汇总:       {len(emp_summary)} 人")
    print("=" * 60)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n[ERR] {e}")
        import traceback
        traceback.print_exc()
        if sys.stdin.isatty():
            input("\n按 Enter 键退出...")
        sys.exit(1)
